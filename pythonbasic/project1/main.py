import openpyxl as opx

inv_file = opx.load_workbook("C:\\growBalagrow\\pythonbasic\\project1\\inventory.xlsx")
product_list = inv_file["Sheet1"]

num_products_per_supplier = {}
total_value_per_supplier = {}
products_under_ten = {}


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory= product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    total_price = product_list.cell(product_row, 5)
    
    # Calculate number of products supplied by each company
    if supplier_name in num_products_per_supplier:
        current_num_products = num_products_per_supplier[supplier_name]
        num_products_per_supplier[supplier_name] = current_num_products + 1
    else:
        num_products_per_supplier[supplier_name] = 1
     
    # Calculate total price of the inventory per supplier    
    if supplier_name in total_value_per_supplier:
        existing_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = existing_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price
    
    # find product that inventory less than 10
    if inventory < 10:
        products_under_ten[int(product_num)] = int(inventory)
    
    # add total price value into new column
    total_price.value = inventory * price


print(num_products_per_supplier)
print(total_value_per_supplier)
print(products_under_ten)
    
inv_file.save("C:\\growBalagrow\\pythonbasic\\project1\\inventory_with_value.xlsx")
    
    
        





